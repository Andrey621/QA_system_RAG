#!/usr/bin/env python
"""
Скрипт-утилита для отправки датасетов на дообучение через GraphRAG API

Поддерживает источники данных в формате CSV или Excel (xlsx/xlsm) с колонками:
    - instruction (обязательна)
    - response (обязательна)
    - context (необязательна)

Пример использования:
    python scripts/submit_fine_tune.py data/examples.csv \
        --base-url http://localhost:8000 \
        --user-id 42 \
        --run-name support-bot-v1 \
        --mode async --poll
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

try:
    import requests  # type: ignore
except ImportError:  # pragma: no cover
    requests = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

from jsonl_dataset import load_examples_from_jsonl, clean_value as _clean_value

REQUIRED_COLUMNS = ("instruction", "response")
OPTIONAL_COLUMNS = ("context",)


def _normalize_row(row: Dict[str, object]) -> Dict[str, str]:
    return {str(key).strip().lower(): _clean_value(value) for key, value in row.items() if key}


def _extract_example(
    row: Dict[str, str],
    instruction_col: str,
    response_col: str,
    context_col: Optional[str],
) -> Optional[Dict[str, str]]:
    instruction = row.get(instruction_col, "")
    response = row.get(response_col, "")
    if not instruction or not response:
        return None
    example: Dict[str, str] = {
        "instruction": instruction,
        "response": response,
    }
    if context_col:
        context_val = row.get(context_col, "")
        if context_val:
            example["context"] = context_val
    return example


def _read_csv(path: Path) -> List[Dict[str, object]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV файл не содержит заголовок.")
        return [dict(row) for row in reader]


def _read_excel(path: Path) -> List[Dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("Для чтения Excel установите пакет openpyxl.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(_clean_value(cell) or "") for cell in rows[0]]
    data_rows: List[Dict[str, object]] = []
    for row in rows[1:]:
        item = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        data_rows.append(item)
    return data_rows


def load_examples_from_tabular(
    file_path: Path,
    *,
    instruction_column: str,
    response_column: str,
    context_column: Optional[str],
    limit: Optional[int],
) -> List[Dict[str, str]]:
    if file_path.suffix.lower() == ".csv":
        rows = _read_csv(file_path)
    else:
        rows = _read_excel(file_path)
    examples: List[Dict[str, str]] = []
    for raw_row in rows:
        normalized_row = _normalize_row(raw_row)
        example = _extract_example(
            normalized_row,
            instruction_col=instruction_column.lower(),
            response_col=response_column.lower(),
            context_col=context_column.lower() if context_column else None,
        )
        if example:
            examples.append(example)
            if limit and len(examples) >= limit:
                break
    if not examples:
        raise ValueError("Не найдено ни одной валидной строки с полями instruction/response.")
    return examples


def load_examples(args: argparse.Namespace) -> List[Dict[str, str]]:
    file_path: Path = args.file
    if not file_path.exists():
        raise FileNotFoundError(f"Файл {file_path} не найден.")
    suffix = file_path.suffix.lower()
    if suffix == ".jsonl":
        return load_examples_from_jsonl(
            file_path,
            split=args.split,
            prompt_key=args.jsonl_prompt_key,
            response_key=args.jsonl_response_key,
            evidence_key=args.jsonl_evidence_key,
            task_key=args.jsonl_task_key,
            labels_key=args.jsonl_labels_key,
            meta_key=args.jsonl_meta_key,
            split_key=args.jsonl_split_key,
            include_meta=args.include_meta_in_context,
            limit=args.limit,
        )
    if suffix in {".csv", ".xlsx", ".xlsm"}:
        return load_examples_from_tabular(
            file_path,
            instruction_column=args.instruction_column,
            response_column=args.response_column,
            context_column=args.context_column,
            limit=args.limit,
        )
    raise ValueError("Поддерживаются JSONL, CSV или Excel (xlsx/xlsm) файлы.")


def build_payload(args: argparse.Namespace, examples: List[Dict[str, str]]) -> Dict[str, object]:
    payload: Dict[str, object] = {
        "user_id": args.user_id,
        "examples": examples,
        "base_model_name": args.base_model_name,
        "learning_rate": args.learning_rate,
        "num_epochs": args.num_epochs,
        "max_seq_length": args.max_seq_length,
        "gradient_accumulation_steps": args.gradient_accumulation_steps,
        "trust_remote_code": not args.no_trust_remote_code,
    }
    if args.output_dir:
        payload["output_dir"] = args.output_dir
    if args.run_name:
        payload["run_name"] = args.run_name
    return payload


def submit_request(base_url: str, endpoint: str, payload: Dict[str, object], timeout: float) -> Dict[str, object]:
    if requests is None:
        raise RuntimeError("Для отправки запросов установите пакет requests.")
    url = f"{base_url.rstrip('/')}{endpoint}"
    response = requests.post(url, json=payload, timeout=timeout)
    response.raise_for_status()
    return response.json()


def poll_job_status(
    base_url: str,
    job_id: str,
    *,
    timeout: float,
    poll_interval: float,
) -> Dict[str, object]:
    status_url = f"{base_url.rstrip('/')}/fine_tune_model/status/{job_id}"
    print(f"[poll] Следим за job_id={job_id} (интервал {poll_interval}s)")
    while True:
        response = requests.get(status_url, timeout=timeout)
        response.raise_for_status()
        data = response.json()
        status = data.get("status")
        message = data.get("message", "")
        print(f"[poll] status={status} message={message}")
        if status in {"completed", "failed"}:
            return data
        time.sleep(poll_interval)


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Submit fine-tuning dataset to GraphRAG API.")
    parser.add_argument("file", type=Path, help="Путь к JSONL/CSV/Excel файлу с примерами.")
    parser.add_argument("--base-url", default="http://localhost:8000", help="Базовый URL FastAPI сервера.")
    parser.add_argument("--mode", choices=("sync", "async"), default="sync",
                        help="Режим запуска дообучения (sync = дождаться, async = очередь).")
    parser.add_argument("--poll", action="store_true", help="При async режиме — ожидать завершения задачи.")
    parser.add_argument("--poll-interval", type=float, default=10.0,
                        help="Интервал (сек) между запросами статуса при polling.")
    parser.add_argument("--user-id", type=int, required=True, help="Идентификатор пользователя/арендатора.")
    parser.add_argument("--base-model-name", default="Qwen/Qwen1.5-0.5B-Chat",
                        help="HuggingFace модель, которую дообучаем.")
    parser.add_argument("--learning-rate", type=float, default=5e-5, help="Learning rate.")
    parser.add_argument("--num-epochs", type=int, default=1, help="Количество эпох.")
    parser.add_argument("--max-seq-length", type=int, default=512, help="Максимальная длина последовательности.")
    parser.add_argument("--gradient-accumulation-steps", type=int, default=1,
                        help="Gradient accumulation steps.")
    parser.add_argument("--output-dir", help="Папка, куда сохранять результаты дообучения.")
    parser.add_argument("--run-name", help="Имя запуска (для логов/структуры каталогов).")
    parser.add_argument("--split", help="Фильтр по значению split внутри JSONL (например, train/test).")
    parser.add_argument("--limit", type=int, help="Максимальное количество примеров, загружаемых из файла.")
    parser.add_argument("--instruction-column", default="instruction",
                        help="Имя столбца с инструкцией.")
    parser.add_argument("--response-column", default="response",
                        help="Имя столбца с ответом.")
    parser.add_argument("--context-column", default="context",
                        help="Имя столбца с контекстом (опционально).")
    parser.add_argument("--jsonl-prompt-key", default="prompt", help="Ключ с инструкцией в JSONL.")
    parser.add_argument("--jsonl-response-key", default="response", help="Ключ с ответом в JSONL.")
    parser.add_argument("--jsonl-evidence-key", default="evidence",
                        help="Ключ со списком evidence (доб. в контекст).")
    parser.add_argument("--jsonl-task-key", default="task", help="Ключ с типом задачи.")
    parser.add_argument("--jsonl-labels-key", default="labels", help="Ключ со списком меток.")
    parser.add_argument("--jsonl-meta-key", default="meta", help="Ключ с метаданными.")
    parser.add_argument("--jsonl-split-key", default="split", help="Ключ с названием сплита.")
    parser.add_argument("--include-meta-in-context", action="store_true",
                        help="Добавлять meta из JSONL в текст контекста.")
    parser.add_argument("--no-trust-remote-code", action="store_true",
                        help="Отключить trust_remote_code при загрузке модели.")
    parser.add_argument("--timeout", type=float, default=30.0,
                        help="HTTP таймаут для запросов к API.")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        examples = load_examples(args)
    except Exception as exc:
        print(f"[error] Не удалось загрузить файл: {exc}", file=sys.stderr)
        return 1

    payload = build_payload(args, examples)
    endpoint = "/fine_tune_model/" if args.mode == "sync" else "/fine_tune_model/async"

    try:
        response = submit_request(args.base_url, endpoint, payload, args.timeout)
    except requests.HTTPError as exc:
        detail = exc.response.text if exc.response is not None else str(exc)
        print(f"[error] Запрос вернул ошибку: {detail}", file=sys.stderr)
        return 1
    except Exception as exc:  # pragma: no cover
        print(f"[error] Не удалось отправить запрос: {exc}", file=sys.stderr)
        return 1

    print("[info] Ответ сервера:")
    print(json.dumps(response, ensure_ascii=False, indent=2))

    if args.mode == "async" and args.poll:
        job_id = response.get("job_id")
        if not job_id:
            print("[warn] Сервер не вернул job_id, polling невозможен.")
            return 0
        final_status = poll_job_status(
            args.base_url,
            job_id,
            timeout=args.timeout,
            poll_interval=args.poll_interval,
        )
        print("[info] Итоговый статус:")
        print(json.dumps(final_status, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
